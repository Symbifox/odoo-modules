import base64
import io
import logging
from collections import defaultdict
from datetime import datetime

from markupsafe import escape as html_escape

from odoo import api, fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger.warning("openpyxl not installed — Excel generation disabled")
    openpyxl = None


class HourBankClient(models.Model):
    _name = 'hour.bank.client'
    _description = "Configuration de banque d'heures client"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'name'

    name = fields.Char(
        compute='_compute_name', store=True, string="Nom",
    )
    partner_id = fields.Many2one(
        'res.partner', required=True, string="Client",
        tracking=True,
    )
    project_ids = fields.Many2many(
        'project.project', string="Projets inclus",
    )
    product_filter_mode = fields.Selection([
        ('all', 'Toutes les lignes de facture'),
        ('include', 'Inclure seulement ces produits'),
        ('exclude', 'Exclure ces produits'),
    ], string="Filtre produits", default='all', required=True)
    filter_product_ids = fields.Many2many(
        'product.product', string="Produits",
    )
    company_filter_mode = fields.Selection([
        ('all', 'Toutes les sociétés'),
        ('include', 'Inclure seulement ces sociétés'),
        ('exclude', 'Exclure ces sociétés'),
    ], string="Filtre sociétés", default='all', required=True)
    filter_company_ids = fields.Many2many(
        'res.company', string="Sociétés",
    )
    invoice_partner_ids = fields.Many2many(
        'res.partner', 'hour_bank_invoice_partner_rel',
        string="Partenaires de facturation",
        help="Si renseigné, seules les factures émises à ces partenaires seront comptées. "
             "Sinon, toutes les factures du partenaire commercial seront incluses.",
    )
    extra_invoice_ids = fields.Many2many(
        'account.move', string="Factures additionnelles",
        domain="[('move_type', '=', 'out_invoice'), ('state', '=', 'posted')]",
    )
    adjustment_ids = fields.One2many(
        'hour.bank.adjustment', 'hour_bank_id',
        string="Ajustements manuels",
    )
    recipient_ids = fields.Many2many(
        'res.partner', 'hour_bank_recipient_rel',
        string="Destinataires du rapport",
    )
    active = fields.Boolean(default=True)
    company_id = fields.Many2one(
        'res.company', default=lambda self: self.env.company,
    )
    current_balance = fields.Float(
        compute='_compute_balance', string="Solde actuel",
    )
    last_report_date = fields.Datetime(
        string="Dernier rapport", tracking=True,
    )
    auto_send = fields.Boolean(
        string="Envoi automatique", default=False, tracking=True,
    )
    send_frequency = fields.Selection([
        ('weekly', 'Hebdomadaire (lundi)'),
        ('biweekly', 'Aux deux semaines (lundi)'),
        ('monthly', 'Mensuel (1er du mois)'),
    ], string="Fréquence d'envoi", default='weekly')

    # ---- Threshold notifications ----
    threshold_mode = fields.Selection([
        ('disabled', 'Désactivé'),
        ('unbilled', 'Heures non facturées'),
        ('budget_pct', '% du budget alloué'),
        ('balance_floor', 'Solde résiduel sous seuil'),
    ], string="Mode de palier", default='disabled', required=True, tracking=True)
    threshold_budget_hours = fields.Float(
        string="Budget alloué (h)", tracking=True,
        help="Budget de référence utilisé en mode '%% du budget alloué'. "
             "Changer cette valeur réarme tous les paliers.",
    )
    threshold_line_ids = fields.One2many(
        'hour.bank.threshold.line', 'bank_id',
        string="Paliers à surveiller",
    )
    threshold_event_ids = fields.One2many(
        'hour.bank.threshold.event', 'bank_id',
        string="Historique des alertes",
    )
    threshold_event_count = fields.Integer(
        compute='_compute_threshold_event_count', string="Alertes envoyées",
    )
    alert_recipient_ids = fields.Many2many(
        'res.partner', 'hour_bank_alert_recipient_rel',
        string="Destinataires des alertes",
        help="Destinataires des alertes de palier. Laisser vide pour réutiliser "
             "les destinataires du rapport périodique.",
    )
    notify_internal_followers = fields.Boolean(
        string="Notifier les followers internes", default=True,
        help="Si coché, poste aussi un message_notify sur le chatter pour "
             "alerter les followers internes Blue Fox.",
    )

    # ------------------------------------------------------------------
    # Computed
    # ------------------------------------------------------------------

    @api.depends('partner_id')
    def _compute_name(self):
        for rec in self:
            rec.name = rec.partner_id.name or _("Nouveau")

    def _compute_balance(self):
        for rec in self:
            data = rec._get_report_data()
            rec.current_balance = data['balance']

    # ------------------------------------------------------------------
    # Core: build report data
    # ------------------------------------------------------------------

    def _get_report_data(self, date_from=None, date_to=None):
        """Build the full hour-bank dataset (debits + credits) for self.

        Returns dict with keys:
            entries: list of dicts sorted date ASC
            balance: final cumulative balance
            summary_by_project: {project_name: total_hours}
            summary_by_month: {(year, month): {project_name: hours}}
            last_update: datetime
        """
        self.ensure_one()

        if not self.project_ids:
            return {
                'entries': [],
                'balance': 0.0,
                'summary_by_project': {},
                'summary_by_month': {},
                'last_update': fields.Datetime.now(),
            }

        project_ids = tuple(self.project_ids.ids)
        partner_id = self.partner_id.commercial_partner_id.id or self.partner_id.id

        # ---- Debits: timesheets ----
        ts_clauses = ["aal.project_id IN %s"]
        ts_params = [project_ids]
        if date_from:
            ts_clauses.append("aal.date >= %s")
            ts_params.append(date_from)
        if date_to:
            ts_clauses.append("aal.date <= %s")
            ts_params.append(date_to)

        ts_sql = """
            SELECT
                aal.date AS entry_date,
                -aal.unit_amount AS hours,
                aal.name AS description,
                pp.name->>'en_US' AS project_name,
                pt.name AS task_name,
                'debit' AS entry_type,
                aal.id AS source_id
            FROM account_analytic_line aal
            JOIN project_project pp ON pp.id = aal.project_id
            LEFT JOIN project_task pt ON pt.id = aal.task_id
            WHERE """ + " AND ".join(ts_clauses)

        # ---- Credits: posted customer invoices ----
        inv_clauses = [
            "am.move_type = 'out_invoice'",
            "am.state = 'posted'",
        ]
        inv_params = []

        # Company filter
        if self.filter_company_ids and self.company_filter_mode == 'include':
            inv_clauses.append("am.company_id IN %s")
            inv_params.append(tuple(self.filter_company_ids.ids))
        elif self.filter_company_ids and self.company_filter_mode == 'exclude':
            inv_clauses.append("am.company_id NOT IN %s")
            inv_params.append(tuple(self.filter_company_ids.ids))

        # Partner filter: specific partners or commercial partner fallback
        extra_ids = tuple(self.extra_invoice_ids.ids) if self.extra_invoice_ids else ()
        if self.invoice_partner_ids:
            partner_filter_ids = tuple(self.invoice_partner_ids.ids)
            if extra_ids:
                inv_clauses.append("(am.partner_id IN %s OR am.id IN %s)")
                inv_params += [partner_filter_ids, extra_ids]
            else:
                inv_clauses.append("am.partner_id IN %s")
                inv_params.append(partner_filter_ids)
        else:
            if extra_ids:
                inv_clauses.append("(am.commercial_partner_id = %s OR am.id IN %s)")
                inv_params += [partner_id, extra_ids]
            else:
                inv_clauses.append("am.commercial_partner_id = %s")
                inv_params.append(partner_id)

        if self.filter_product_ids and self.product_filter_mode == 'include':
            inv_clauses.append("aml.product_id IN %s")
            inv_params.append(tuple(self.filter_product_ids.ids))
        elif self.filter_product_ids and self.product_filter_mode == 'exclude':
            inv_clauses.append("(aml.product_id IS NULL OR aml.product_id NOT IN %s)")
            inv_params.append(tuple(self.filter_product_ids.ids))

        if date_from:
            inv_clauses.append("am.invoice_date >= %s")
            inv_params.append(date_from)
        if date_to:
            inv_clauses.append("am.invoice_date <= %s")
            inv_params.append(date_to)

        inv_clauses += ["aml.quantity > 0", "aml.display_type = 'product'"]

        inv_sql = """
            SELECT
                am.invoice_date AS entry_date,
                aml.quantity AS hours,
                am.name AS description,
                'Heures factur\u00e9es' AS project_name,
                NULL AS task_name,
                'credit' AS entry_type,
                aml.id AS source_id
            FROM account_move_line aml
            JOIN account_move am ON am.id = aml.move_id
            WHERE """ + " AND ".join(inv_clauses)

        # ---- Adjustments: manual entries ----
        adj_clauses = ["hba.hour_bank_id = %s"]
        adj_params = [self.id]
        if date_from:
            adj_clauses.append("hba.date >= %s")
            adj_params.append(date_from)
        if date_to:
            adj_clauses.append("hba.date <= %s")
            adj_params.append(date_to)

        adj_sql = """
            SELECT
                hba.date AS entry_date,
                hba.hours AS hours,
                hba.description AS description,
                'Ajustement' AS project_name,
                NULL AS task_name,
                CASE WHEN hba.hours >= 0 THEN 'credit' ELSE 'debit' END AS entry_type,
                hba.id + 1000000000 AS source_id
            FROM hour_bank_adjustment hba
            WHERE """ + " AND ".join(adj_clauses)

        # ---- Union + order ----
        sql = """
            SELECT * FROM (
                (%s)
                UNION ALL
                (%s)
                UNION ALL
                (%s)
            ) combined
            ORDER BY entry_date ASC, entry_type DESC, source_id ASC
        """ % (ts_sql, inv_sql, adj_sql)
        params = ts_params + inv_params + adj_params
        self.env.cr.execute(sql, params)
        rows = self.env.cr.dictfetchall()

        # ---- Build entries with cumulative balance ----
        entries = []
        balance = 0.0
        summary_by_project = defaultdict(float)
        # summary_by_month: {(year, month): {project: hours}}
        summary_by_month = defaultdict(lambda: defaultdict(float))

        for row in rows:
            balance += row['hours']
            entry = {
                'date': row['entry_date'],
                'hours': row['hours'],
                'cumulative': balance,
                'description': row['description'] or '',
                'project': row['project_name'] or '',
                'task': row['task_name'] or '',
                'entry_type': row['entry_type'],
            }
            entries.append(entry)

            # Summaries
            summary_by_project[entry['project']] += row['hours']
            dt = row['entry_date']
            summary_by_month[(dt.year, dt.month)][entry['project']] += row['hours']

        # Reverse for display (most recent first)
        entries.reverse()

        return {
            'entries': entries,
            'balance': balance,
            'summary_by_project': dict(summary_by_project),
            'summary_by_month': {k: dict(v) for k, v in summary_by_month.items()},
            'last_update': fields.Datetime.now(),
        }

    # ------------------------------------------------------------------
    # PDF generation
    # ------------------------------------------------------------------

    def action_generate_pdf(self):
        """Generate the PDF report and return download action."""
        self.ensure_one()
        return self.env.ref(
            'bf_hour_bank.action_report_hour_bank'
        ).report_action(self)

    def _get_pdf_binary(self):
        """Generate PDF as binary for attachment purposes."""
        self.ensure_one()
        pdf_content, _content_type = self.env['ir.actions.report']._render_qweb_pdf(
            'bf_hour_bank.action_report_hour_bank', self.ids,
        )
        return pdf_content

    # ------------------------------------------------------------------
    # Excel generation
    # ------------------------------------------------------------------

    def action_generate_xlsx(self):
        """Generate Excel and return download action."""
        self.ensure_one()
        xlsx_data = self._generate_xlsx_binary()
        filename = "Banque_heures_%s_%s.xlsx" % (
            self.partner_id.name.replace(' ', '_'),
            fields.Date.today().isoformat(),
        )
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'new',
        }

    def _generate_xlsx_binary(self):
        """Build a 4-sheet .xlsx and return raw bytes."""
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("La biblioth\u00e8que openpyxl n'est pas install\u00e9e."))

        data = self._get_report_data()
        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_hex = (self.env.company.report_brand_dark or '#212529').lstrip('#').upper()
        header_fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type='solid')
        credit_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        number_fmt = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        def safe_cell_value(value):
            """Prevent Excel formula injection by prefixing dangerous characters."""
            if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@'):
                return "'" + value
            return value

        def style_header(ws, cols):
            for idx, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=idx, value=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # ─── Sheet 1: Feuilles de temps ───
        ws1 = wb.active
        ws1.title = "Feuilles de temps"
        cols1 = ["Date", "Nombre d'heures", "Solde cumulatif", "Description", "Projet", "Tâche"]
        style_header(ws1, cols1)

        for i, entry in enumerate(data['entries'], 2):
            ws1.cell(row=i, column=1, value=entry['date']).number_format = 'YYYY-MM-DD'
            ws1.cell(row=i, column=2, value=entry['hours']).number_format = number_fmt
            ws1.cell(row=i, column=3, value=entry['cumulative']).number_format = number_fmt
            ws1.cell(row=i, column=4, value=safe_cell_value(entry['description']))
            ws1.cell(row=i, column=5, value=safe_cell_value(entry['project']))
            ws1.cell(row=i, column=6, value=safe_cell_value(entry['task']))
            for c in range(1, 7):
                ws1.cell(row=i, column=c).border = thin_border
            if entry['entry_type'] == 'credit':
                for c in range(1, 7):
                    ws1.cell(row=i, column=c).fill = credit_fill

        # Auto-width
        for col_idx in range(1, 7):
            ws1.column_dimensions[get_column_letter(col_idx)].width = (
                [14, 16, 16, 40, 25, 25][col_idx - 1]
            )

        # ─── Sheet 2: Sommaire par projet ───
        ws2 = wb.create_sheet("Sommaire par projet")
        cols2 = ["Projet", "Heures totales"]
        style_header(ws2, cols2)
        row = 2
        for project, hours in sorted(data['summary_by_project'].items()):
            ws2.cell(row=row, column=1, value=project).border = thin_border
            cell = ws2.cell(row=row, column=2, value=hours)
            cell.number_format = number_fmt
            cell.border = thin_border
            row += 1
        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 18

        # ─── Sheet 3: Synth\u00e8se par mois ───
        ws3 = wb.create_sheet("Synth\u00e8se par mois")
        # Collect all projects for column headers
        all_projects = sorted(set(
            p for month_data in data['summary_by_month'].values()
            for p in month_data
        ))
        cols3 = ["Mois"] + all_projects + ["Total"]
        style_header(ws3, cols3)

        row = 2
        for (year, month), proj_hours in sorted(data['summary_by_month'].items()):
            ws3.cell(row=row, column=1, value=f"{year}-{month:02d}").border = thin_border
            month_total = 0.0
            for col_idx, proj in enumerate(all_projects, 2):
                val = proj_hours.get(proj, 0.0)
                cell = ws3.cell(row=row, column=col_idx, value=val)
                cell.number_format = number_fmt
                cell.border = thin_border
                month_total += val
            cell = ws3.cell(row=row, column=len(all_projects) + 2, value=month_total)
            cell.number_format = number_fmt
            cell.border = thin_border
            cell.font = Font(bold=True)
            row += 1

        ws3.column_dimensions['A'].width = 14
        for idx in range(2, len(all_projects) + 3):
            ws3.column_dimensions[get_column_letter(idx)].width = 20

        # ─── Sheet 4: \u00c0 facturer ───
        ws4 = wb.create_sheet("\u00c0 facturer")
        cols4 = ["Date", "Heures", "Description", "Projet", "T\u00e2che"]
        style_header(ws4, cols4)

        # Entries since last credit (invoice)
        unbilled = []
        for entry in reversed(data['entries']):
            if entry['entry_type'] == 'credit':
                break
            unbilled.insert(0, entry)

        row = 2
        total_unbilled = 0.0
        for entry in unbilled:
            ws4.cell(row=row, column=1, value=entry['date']).number_format = 'YYYY-MM-DD'
            ws4.cell(row=row, column=2, value=entry['hours']).number_format = number_fmt
            ws4.cell(row=row, column=3, value=safe_cell_value(entry['description']))
            ws4.cell(row=row, column=4, value=safe_cell_value(entry['project']))
            ws4.cell(row=row, column=5, value=safe_cell_value(entry['task']))
            for c in range(1, 6):
                ws4.cell(row=row, column=c).border = thin_border
            total_unbilled += entry['hours']
            row += 1

        # Total row
        ws4.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws4.cell(row=row, column=1).border = thin_border
        total_cell = ws4.cell(row=row, column=2, value=total_unbilled)
        total_cell.number_format = number_fmt
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border

        ws4.column_dimensions['A'].width = 14
        ws4.column_dimensions['B'].width = 14
        ws4.column_dimensions['C'].width = 40
        ws4.column_dimensions['D'].width = 25
        ws4.column_dimensions['E'].width = 25

        # Write to buffer
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Open send wizard
    # ------------------------------------------------------------------

    def action_open_send_wizard(self):
        """Open the email-send wizard with pre-generated attachments."""
        self.ensure_one()
        wizard = self.env['hour.bank.send.wizard'].create({
            'hour_bank_id': self.id,
        })
        return {
            'type': 'ir.actions.act_window',
            'name': _("Envoyer le rapport"),
            'res_model': 'hour.bank.send.wizard',
            'res_id': wizard.id,
            'views': [[False, 'form']],
            'target': 'new',
        }

    # ------------------------------------------------------------------
    # Automated sending (cron)
    # ------------------------------------------------------------------

    def _send_report_to_recipients(self):
        """Generate and send branded report email to recipients.

        Used by both the cron and can be called programmatically.
        """
        self.ensure_one()
        if not self.recipient_ids:
            _logger.info("Hour bank %s: no recipients configured, skipping", self.name)
            return

        WizardModel = self.env['hour.bank.send.wizard']

        partner_name = self.partner_id.name.replace(' ', '_')
        date_str = fields.Date.today().isoformat()

        # Generate PDF
        pdf_data = self._get_pdf_binary()
        pdf_att = self.env['ir.attachment'].create({
            'name': "Banque_heures_%s_%s.pdf" % (partner_name, date_str),
            'type': 'binary',
            'datas': base64.b64encode(pdf_data),
            'mimetype': 'application/pdf',
            'res_model': self._name,
            'res_id': self.id,
        })

        # Generate Excel
        xlsx_data = self._generate_xlsx_binary()
        xlsx_att = self.env['ir.attachment'].create({
            'name': "Banque_heures_%s_%s.xlsx" % (partner_name, date_str),
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        # Build branded body
        inner_body = (
            '<p style="font-size:16px;line-height:26px;color:#374151;'
            'margin:0 0 16px 0;">Bonjour,</p>'
            '<p style="font-size:16px;line-height:26px;color:#374151;'
            'margin:0 0 20px 0;">Veuillez trouver ci-joint '
            "l'&#233;tat des banques d'heures pour "
            "<strong>%s</strong>.</p>"  # noqa: S001
            '<p style="font-size:16px;line-height:26px;color:#374151;'
            'margin:0 0 20px 0;">Les fichiers PDF et Excel sont '
            "en pi&#232;ces jointes.</p>"
            '<p style="font-size:16px;line-height:26px;color:#374151;'
            'margin:0;">Cordialement,<br/>Blue Fox</p>'
        ) % html_escape(self.partner_id.name)

        body_html = WizardModel.new({})._wrap_branded_body(inner_body)

        subject = "État des banques d'heures — %s" % self.partner_id.name
        sender = self.company_id.email or self.env.user.email_formatted

        for partner in self.recipient_ids:
            if not partner.email:
                continue
            mail = self.env['mail.mail'].sudo().create({
                'subject': subject,
                'body_html': body_html,
                'email_from': sender,
                # recipient_ids ONLY — never also set email_to to the same
                # person: Odoo's _prepare_outgoing_list builds one outgoing
                # email for email_to AND one per recipient_ids partner (no
                # dedup), so the recipient gets two identical SMTP deliveries.
                'recipient_ids': [(4, partner.id)],
                'attachment_ids': [(6, 0, [pdf_att.id, xlsx_att.id])],
            })
            mail.send()

        self.write({'last_report_date': fields.Datetime.now()})

        recipient_names = ', '.join(self.recipient_ids.mapped('name'))
        self.message_post(
            body=_("Rapport automatique envoyé à %s") % recipient_names,
            message_type='comment',
            subtype_xmlid='mail.mt_note',
            attachment_ids=[pdf_att.id, xlsx_att.id],
        )
        _logger.info("Hour bank %s: report sent to %s", self.name, recipient_names)

    @api.model
    def _cron_send_reports(self):
        """Cron job: send reports for all active auto-send hour banks."""
        today = fields.Date.today()
        weekday = today.weekday()  # 0=Monday
        day = today.day

        banks = self.search([('auto_send', '=', True), ('active', '=', True)])
        for bank in banks:
            try:
                send = False
                if bank.send_frequency == 'weekly' and weekday == 0:
                    send = True
                elif bank.send_frequency == 'biweekly' and weekday == 0:
                    # Send on even ISO weeks
                    if today.isocalendar()[1] % 2 == 0:
                        send = True
                elif bank.send_frequency == 'monthly' and day == 1:
                    send = True

                if send:
                    bank._send_report_to_recipients()
            except Exception:
                _logger.exception(
                    "Hour bank cron: failed to send report for %s", bank.name
                )

    # ------------------------------------------------------------------
    # Threshold notifications
    # ------------------------------------------------------------------

    @api.depends('threshold_event_ids')
    def _compute_threshold_event_count(self):
        for rec in self:
            rec.threshold_event_count = len(rec.threshold_event_ids)

    def _get_alert_recipients(self):
        """Return partners to email for threshold alerts.

        Fallback: if alert_recipient_ids is empty, reuse recipient_ids
        (same audience as the periodic report). Drops partners without email.
        """
        self.ensure_one()
        partners = self.alert_recipient_ids or self.recipient_ids
        return partners.filtered(lambda p: p.email)

    def _last_posted_invoice_id(self):
        """Return id (int) of the most recently posted invoice that would
        be counted by this bank's filters, or 0 if none.

        Used as the period_key for the 'unbilled' mode: when a new invoice
        is posted, the key changes and thresholds re-arm.
        """
        self.ensure_one()
        partner_id = self.partner_id.commercial_partner_id.id or self.partner_id.id
        domain = [
            ('move_type', '=', 'out_invoice'),
            ('state', '=', 'posted'),
        ]
        if self.invoice_partner_ids:
            domain.append(('partner_id', 'in', self.invoice_partner_ids.ids))
        else:
            domain.append(('commercial_partner_id', '=', partner_id))
        if self.filter_company_ids and self.company_filter_mode == 'include':
            domain.append(('company_id', 'in', self.filter_company_ids.ids))
        elif self.filter_company_ids and self.company_filter_mode == 'exclude':
            domain.append(('company_id', 'not in', self.filter_company_ids.ids))
        inv = self.env['account.move'].sudo().search(
            domain, order='invoice_date desc, id desc', limit=1,
        )
        # Also include extra_invoice_ids — if the most recent extra invoice
        # is newer than the domain match, use that one.
        if self.extra_invoice_ids:
            latest_extra = self.extra_invoice_ids.sorted(
                lambda m: (m.invoice_date or fields.Date.from_string('1900-01-01'), m.id),
                reverse=True,
            )[:1]
            if latest_extra and (
                not inv
                or (latest_extra.invoice_date or fields.Date.from_string('1900-01-01'))
                > (inv.invoice_date or fields.Date.from_string('1900-01-01'))
            ):
                inv = latest_extra
        return inv.id if inv else 0

    def _current_period_key(self):
        """Return the period_key string for the current mode.

        - unbilled    -> str(id of most-recent posted invoice), '0' if none
        - budget_pct  -> str(threshold_budget_hours)  (changes when budget changes)
        - balance_floor -> '' (handled via crossing logic in _check_thresholds)
        - disabled    -> ''
        """
        self.ensure_one()
        if self.threshold_mode == 'unbilled':
            return str(self._last_posted_invoice_id())
        if self.threshold_mode == 'budget_pct':
            return '%.4f' % (self.threshold_budget_hours or 0.0)
        return ''

    def _compute_threshold_measurements(self, report_data=None):
        """Return a dict {current_unbilled, current_balance, current_pct,
        current_period_key, last_invoice_date}.

        Computes once per call by reusing _get_report_data() output if provided.
        """
        self.ensure_one()
        data = report_data if report_data is not None else self._get_report_data()
        balance = data['balance']

        # current_unbilled: sum of debit hours since the last credit entry.
        # Entries are reversed (most recent first), so iterate until we hit a credit.
        unbilled = 0.0
        last_invoice_date = None
        for entry in data['entries']:
            if entry['entry_type'] == 'credit':
                last_invoice_date = entry['date']
                break
            # debit hours are negative; "unbilled hours consumed" is the absolute value
            unbilled += -entry['hours']

        budget = self.threshold_budget_hours or 0.0
        pct = (unbilled / budget * 100.0) if budget > 0 else 0.0

        return {
            'current_unbilled': unbilled,
            'current_balance': balance,
            'current_pct': pct,
            'current_period_key': self._current_period_key(),
            'last_invoice_date': last_invoice_date,
        }

    def _check_thresholds(self):
        """Evaluate every active threshold line on self and fire alerts as needed.

        Returns the number of alerts fired (useful for tests and UI feedback).
        """
        self.ensure_one()
        if self.threshold_mode == 'disabled' or not self.threshold_line_ids:
            return 0
        if not self.active:
            return 0

        data = self._get_report_data()
        m = self._compute_threshold_measurements(report_data=data)
        fired = 0

        for line in self.threshold_line_ids.filtered('active'):
            try:
                should_fire, new_period_key, measured = self._evaluate_threshold(line, m)
            except Exception:
                _logger.exception(
                    "Hour bank %s: threshold evaluation failed for line %s",
                    self.name, line.value,
                )
                continue
            if should_fire:
                try:
                    self._fire_threshold_alert(line, m, measured, new_period_key, report_data=data)
                    fired += 1
                except Exception:
                    _logger.exception(
                        "Hour bank %s: failed to send threshold alert for line %s",
                        self.name, line.value,
                    )
            elif new_period_key is not None and new_period_key != (line.last_fired_period_key or ''):
                # Silent re-arm: update period_key without firing
                # (used by balance_floor when balance rises back above hysteresis)
                line.write({'last_fired_period_key': new_period_key})
        return fired

    def _evaluate_threshold(self, line, measurement):
        """Decide whether `line` should fire given current `measurement`.

        Returns tuple (should_fire: bool, new_period_key: str|None, measured: float).
        new_period_key is the value to persist on line.last_fired_period_key:
        - on fire: the new key (consumed)
        - on silent re-arm (balance_floor): the new "armed" key
        - None: no change to persist
        """
        self.ensure_one()
        mode = self.threshold_mode

        if mode == 'unbilled':
            measured = measurement['current_unbilled']
            current_key = measurement['current_period_key']
            rearmed = (line.last_fired_period_key or '') != current_key
            if rearmed and measured >= line.value:
                return True, current_key, measured
            return False, None, measured

        if mode == 'budget_pct':
            measured = measurement['current_pct']
            current_key = measurement['current_period_key']
            rearmed = (line.last_fired_period_key or '') != current_key
            if rearmed and measured >= line.value:
                return True, current_key, measured
            return False, None, measured

        if mode == 'balance_floor':
            measured = measurement['current_balance']
            # Hysteresis: re-arm silently when balance rises above value + 0.5h
            hysteresis = 0.5
            currently_armed = not (line.last_fired_period_key or '')
            if not currently_armed and measured > (line.value + hysteresis):
                # Silent re-arm
                return False, '', measured
            if currently_armed and measured <= line.value:
                # Fire and disarm
                new_key = fields.Datetime.to_string(fields.Datetime.now())
                return True, new_key, measured
            return False, None, measured

        return False, None, 0.0

    def _fire_threshold_alert(self, line, measurement, measured, period_key, report_data=None):
        """Send the alert email + record the event + post on chatter."""
        self.ensure_one()
        WizardModel = self.env['hour.bank.send.wizard']
        partners = self._get_alert_recipients()

        if not partners:
            _logger.info(
                "Hour bank %s: threshold %s reached but no recipients with email",
                self.name, line.name,
            )
            # Still record the event and update line so we don't loop
            line.write({
                'last_fired_date': fields.Datetime.now(),
                'last_fired_period_key': period_key,
            })
            self.env['hour.bank.threshold.event'].create({
                'bank_id': self.id,
                'threshold_line_id': line.id,
                'mode_snapshot': self.threshold_mode,
                'value_snapshot': line.value,
                'measured_value': measured,
                'period_key': period_key,
                'recipient_ids': [(6, 0, [])],
            })
            return

        # Generate XLSX attachment (reuse data if available)
        if report_data is None:
            report_data = self._get_report_data()
        xlsx_data = self._generate_xlsx_binary()
        partner_name = self.partner_id.name.replace(' ', '_')
        date_str = fields.Date.today().isoformat()
        xlsx_att = self.env['ir.attachment'].create({
            'name': "Banque_heures_%s_%s.xlsx" % (partner_name, date_str),
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        # Build branded HTML body
        if self.threshold_mode == 'budget_pct':
            measure_label = "%.1f %% du budget consommés (%.2fh sur %.2fh)" % (
                measured, measurement['current_unbilled'], self.threshold_budget_hours,
            )
        elif self.threshold_mode == 'unbilled':
            measure_label = "%.2fh non facturées" % measured
        else:
            measure_label = "Solde courant : %.2fh" % measured

        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url', '')
        bank_url = "%s/odoo/action-base.action_client/?model=hour.bank.client&view_type=form&id=%d" % (
            base_url.rstrip('/'), self.id,
        ) if base_url else ''

        inner_body = (
            '<p style="font-size:16px;line-height:26px;color:#374151;'
            'margin:0 0 16px 0;">Bonjour,</p>'
            '<p style="font-size:16px;line-height:26px;color:#374151;'
            'margin:0 0 16px 0;">Votre banque d\'heures pour '
            "<strong>%s</strong> vient d'atteindre le palier "
            "<strong>« %s »</strong>.</p>"
            '<table cellpadding="6" cellspacing="0" border="0" '
            'style="border-collapse:collapse;margin:0 0 20px 0;'
            'font-size:14px;color:#374151;">'
            '<tr><td style="color:#6B7280;">Mesure actuelle :</td>'
            '<td><strong>%s</strong></td></tr>'
            '<tr><td style="color:#6B7280;">Solde courant :</td>'
            '<td><strong>%.2fh</strong></td></tr>'
            '</table>'
            '<p style="font-size:16px;line-height:26px;color:#374151;'
            'margin:0 0 20px 0;">Le détail des entrées est joint en '
            'format Excel.</p>'
            '%s'
            '<p style="font-size:16px;line-height:26px;color:#374151;'
            'margin:0;">Cordialement,<br/>Blue Fox</p>'
        ) % (
            html_escape(self.partner_id.name),
            html_escape(line.name or ''),
            html_escape(measure_label),
            measurement['current_balance'],
            ('<p style="font-size:14px;line-height:22px;margin:0 0 20px 0;">'
             '<a href="%s" style="color:#29ABE2;">Ouvrir la banque dans Odoo</a></p>'
             % html_escape(bank_url)) if bank_url else '',
        )

        body_html = WizardModel.new({})._wrap_branded_body(inner_body)
        subject = "Banque d'heures %s — palier « %s » atteint" % (
            self.partner_id.name, line.name or '',
        )
        sender = self.company_id.email or self.env.user.email_formatted

        for partner in partners:
            mail = self.env['mail.mail'].sudo().create({
                'subject': subject,
                'body_html': body_html,
                'email_from': sender,
                # recipient_ids ONLY — see note in _send_report_to_recipients:
                # setting email_to to the same partner duplicates the SMTP send.
                'recipient_ids': [(4, partner.id)],
                'attachment_ids': [(6, 0, [xlsx_att.id])],
            })
            mail.send()

        # Update the line state (consume the period)
        line.write({
            'last_fired_date': fields.Datetime.now(),
            'last_fired_period_key': period_key,
        })

        # Chatter note (audit) — captures the mail.message id for the event
        recipient_names = ', '.join(partners.mapped('name'))
        message = self.message_post(
            body=_("Alerte de palier « %(label)s » envoyée à %(names)s — %(measure)s") % {
                'label': line.name or '',
                'names': recipient_names,
                'measure': measure_label,
            },
            message_type='comment',
            subtype_xmlid='mail.mt_note',
            attachment_ids=[xlsx_att.id],
        )

        # Optional message_notify to internal followers
        if self.notify_internal_followers:
            internal_partners = self.message_partner_ids.filtered(
                lambda p: p.user_ids and not p.user_ids.share
            ) - partners
            if internal_partners:
                try:
                    self.message_notify(
                        partner_ids=internal_partners.ids,
                        subject=subject,
                        body=_(
                            "Alerte de palier « %(label)s » sur %(bank)s. "
                            "Mesure : %(measure)s."
                        ) % {
                            'label': line.name or '',
                            'bank': self.name,
                            'measure': measure_label,
                        },
                    )
                except Exception:
                    _logger.exception(
                        "Hour bank %s: message_notify to internal followers failed",
                        self.name,
                    )

        # Append-only event record
        self.env['hour.bank.threshold.event'].create({
            'bank_id': self.id,
            'threshold_line_id': line.id,
            'mode_snapshot': self.threshold_mode,
            'value_snapshot': line.value,
            'measured_value': measured,
            'period_key': period_key,
            'recipient_ids': [(6, 0, partners.ids)],
            'attachment_id': xlsx_att.id,
            'mail_message_id': message.id if message else False,
        })
        _logger.info(
            "Hour bank %s: threshold %s fired (measured=%.2f, recipients=%s)",
            self.name, line.name, measured, recipient_names,
        )

    def action_check_thresholds_now(self):
        """Manual button: evaluate thresholds and show how many fired."""
        self.ensure_one()
        fired = self._check_thresholds()
        if fired:
            msg = _("%(n)d alerte(s) de palier envoyée(s).", n=fired)
        else:
            msg = _("Aucun palier déclenché. (Tous déjà notifiés ou en-dessous des seuils.)")
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Vérification des paliers"),
                'message': msg,
                'sticky': False,
                'type': 'success' if fired else 'info',
            },
        }

    def action_view_threshold_events(self):
        """Smart button: open event history for this bank."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _("Alertes envoyées"),
            'res_model': 'hour.bank.threshold.event',
            'view_mode': 'list,form',
            'domain': [('bank_id', '=', self.id)],
            'context': {'default_bank_id': self.id},
        }

    @api.model
    def _cron_check_thresholds(self):
        """Cron: evaluate thresholds for all active banks with mode != disabled."""
        banks = self.search([
            ('active', '=', True),
            ('threshold_mode', '!=', 'disabled'),
        ])
        for bank in banks:
            try:
                bank._check_thresholds()
            except Exception:
                _logger.exception(
                    "Hour bank cron: threshold check failed for %s", bank.name,
                )
