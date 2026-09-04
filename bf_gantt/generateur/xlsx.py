# -*- coding: utf-8 -*-
"""La feuille de calcul, dans les deux sens.

Le classeur qui sort est celui qui rentre : mêmes colonnes, mêmes en-têtes, même
ordre. C'est la condition pour qu'un aller-retour soit un vrai aller-retour et
non deux formats qui se ressemblent.

À droite du tableau, une grille de cases colorées redessine l'échéancier. Elle
est décorative : l'import ne la lit jamais, il ne lit que les colonnes nommées.
"""
import io
from datetime import date, datetime, timedelta

import xlsxwriter

from . import geometrie as geo

# L'ordre fait foi. L'import retrouve chaque colonne par son en-tête, pas par sa
# position, mais l'export les écrit dans cet ordre pour rester lisible.
COLONNES = [
    ("lane", "Couloir", 22),
    ("name", "Ligne", 46),
    ("assignee", "Responsable", 18),
    ("start", "Début", 12),
    ("end", "Fin", 12),
    ("days", "Jours", 7),
    ("progress", "Avancement %", 13),
    ("status", "Statut", 13),
    ("milestone", "Jalon", 7),
    ("allocated_hours", "Heures prévues", 14),
    ("effective_hours", "Heures faites", 13),
    ("depends_on", "Précédée par", 26),
    ("ref", "Référence", 14),
]

LIBELLES_STATUT = {
    "done": "Terminé",
    "canceled": "Annulé",
    "overdue": "En retard",
    "in_progress": "En cours",
    "upcoming": "À venir",
}
STATUT_DEPUIS_LIBELLE = {v.lower(): k for k, v in LIBELLES_STATUT.items()}

# Largeur de la grille visuelle, en nombre de colonnes.
GRILLE_MAX = 60

# ⚠️ Un classeur téléversé est un fichier étranger. `read_only=True` évite de le
# charger en entier, mais rien ne bornait le nombre de rangées lues ni la taille
# du fichier : une feuille d'un million de lignes se lisait jusqu'au bout.
LIGNES_MAX = 5000
OCTETS_MAX = 25 * 1024 * 1024


def rendre(payload, avec_grille=True):
    """Rend le classeur en octets."""
    tampon = io.BytesIO()
    livre = xlsxwriter.Workbook(tampon, {
        "in_memory": True,
        "default_date_format": "yyyy-mm-dd",
        "remove_timezone": True,
    })
    accent = payload.get("company", {}).get("color") or "#29ABE1"

    f_titre = livre.add_format({
        "bold": True, "font_size": 14, "font_color": "#2D3031", "font_name": "Lexend"})
    f_sous = livre.add_format({
        "font_size": 9, "font_color": "#6B7379", "font_name": "Lexend"})
    f_entete = livre.add_format({
        "bold": True, "font_size": 9, "font_color": "#FFFFFF", "bg_color": accent,
        "align": "left", "valign": "vcenter", "border": 1, "border_color": "#FFFFFF",
        "font_name": "Lexend", "text_wrap": True})
    f_couloir = livre.add_format({
        "bold": True, "font_size": 9, "bg_color": "#F4F6F8", "font_name": "Lexend"})
    f_texte = livre.add_format({"font_size": 9, "font_name": "Lexend"})
    f_date = livre.add_format({
        "font_size": 9, "num_format": "yyyy-mm-dd", "font_name": "Lexend"})
    f_nombre = livre.add_format({
        "font_size": 9, "num_format": "0.##", "font_name": "Lexend"})
    f_pct = livre.add_format({
        "font_size": 9, "num_format": "0\\ %", "font_name": "Lexend"})

    feuille = livre.add_worksheet("Échéancier")
    feuille.set_default_row(15)

    feuille.write(0, 0, payload.get("title") or "Échéancier", f_titre)
    sous = " · ".join(p for p in [
        payload.get("subtitle") or "",
        payload.get("company", {}).get("name") or "",
        "%s au %s" % (payload["range"]["min"], payload["range"]["max"]),
    ] if p)
    feuille.write(1, 0, sous, f_sous)

    # Le logo de la société, au-dessus du tableau. `insert_image` veut un flux,
    # et l'échelle se donne en fraction : on vise 42 pixels de haut, la hauteur
    # de deux lignes, pour qu'il ne pousse pas l'en-tête vers le bas.
    logo = geo.boite_logo(payload.get("company", {}).get("logo"),
                          hauteur_max=42.0, largeur_max=220.0)
    if logo and not logo.get("vectoriel"):
        try:
            import base64
            brut = io.BytesIO(base64.b64decode(logo["b64"]))
            from PIL import Image as _Image
            with _Image.open(io.BytesIO(logo["octets"])) as im:
                l_px, h_px = im.size
            facteur = min(42.0 / h_px, 220.0 / l_px)
            feuille.insert_image(0, len(COLONNES) - 2, "logo.png", {
                "image_data": brut, "x_scale": facteur, "y_scale": facteur,
                "object_position": 3,  # ne bouge ni ne se dimensionne avec les cellules
            })
        except Exception:
            # Un logo illisible ne fait pas tomber le classeur.
            pass

    ligne_entete = 3
    for col, (_cle, libelle, largeur) in enumerate(COLONNES):
        feuille.set_column(col, col, largeur)
        feuille.write(ligne_entete, col, libelle, f_entete)
    feuille.freeze_panes(ligne_entete + 1, 2)
    feuille.autofilter(ligne_entete, 0, ligne_entete, len(COLONNES) - 1)

    # Une référence lisible par tâche, pour la colonne « Précédée par ».
    noms = {t["ref"]: t["name"] for t in payload["tasks"]}
    amont = {}
    for lien in payload.get("deps", []):
        amont.setdefault(lien["to"], []).append(noms.get(lien["from"], lien["from"]))

    couloirs = {c["key"]: c["name"] for c in payload["lanes"]}
    par_couloir = {}
    for tache in payload["tasks"]:
        par_couloir.setdefault(tache["lane"], []).append(tache)

    grille = _grille(payload) if avec_grille else None
    col_grille = len(COLONNES) + 1
    if grille:
        _entete_grille(feuille, livre, grille, ligne_entete, col_grille, accent)

    r = ligne_entete + 1
    for couloir in payload["lanes"]:
        taches = par_couloir.get(couloir["key"], [])
        if not taches:
            continue
        feuille.write(r, 0, couloir["name"], f_couloir)
        for col in range(1, len(COLONNES)):
            feuille.write_blank(r, col, None, f_couloir)
        r += 1
        for tache in sorted(taches, key=lambda t: (t["start"], t["name"] or "")):
            debut = date.fromisoformat(tache["start"])
            fin = date.fromisoformat(tache["end"])
            valeurs = {
                "lane": couloirs.get(tache["lane"], ""),
                "name": tache["name"],
                "assignee": tache.get("assignee", ""),
                "start": debut,
                "end": fin,
                "days": (fin - debut).days + 1,
                "progress": tache.get("progress", 0),
                "status": LIBELLES_STATUT.get(tache["status"], tache["status"]),
                "milestone": "oui" if tache.get("is_milestone") else "",
                "allocated_hours": tache.get("allocated_hours", 0.0),
                "effective_hours": tache.get("effective_hours", 0.0),
                "depends_on": " ; ".join(amont.get(tache["ref"], [])),
                "ref": tache["ref"],
            }
            for col, (cle, _lbl, _w) in enumerate(COLONNES):
                v = valeurs[cle]
                if isinstance(v, date):
                    feuille.write_datetime(r, col, datetime(v.year, v.month, v.day),
                                           f_date)
                elif cle == "progress":
                    feuille.write_number(r, col, v, f_pct)
                elif cle in ("allocated_hours", "effective_hours", "days"):
                    feuille.write_number(r, col, v or 0, f_nombre)
                else:
                    feuille.write(r, col, v, f_texte)
            if grille:
                _barre_grille(feuille, livre, grille, r, col_grille, tache, debut, fin)
            r += 1

    pied = payload.get("company", {}).get("name") or ""
    if payload.get("company", {}).get("tagline"):
        pied = "%s · %s" % (pied, payload["company"]["tagline"])
    feuille.write(r + 1, 0,
                  "Les colonnes nommées font foi. La grille de droite est un dessin : "
                  "l'import ne la lit pas.", f_sous)
    if pied:
        feuille.write(r + 2, 0, pied, f_sous)
    if _plage_reduite(payload):
        feuille.write(r + 3, 0,
                      "⚠️ La plage a été ramenée à %d ans : la grille de droite "
                      "s'arrête avant la fin réelle de certaines lignes. Les "
                      "colonnes Début et Fin, elles, portent les vraies dates."
                      % (geo.PLAGE_MAX_JOURS // 366), f_sous)

    livre.close()
    return tampon.getvalue()


def _plage_reduite(payload):
    debut = date.fromisoformat(payload["range"]["min"])
    fin = date.fromisoformat(payload["range"]["max"])
    return (fin - debut).days > geo.PLAGE_MAX_JOURS


def _grille(payload):
    """Découpe la plage en colonnes de largeur adaptée au nombre de jours.

    ⚠️ Cette fonction lisait `payload["range"]` en direct, donc SANS le plafond
    que `geometrie.construire` applique : elle héritait de l'étendue brute et
    pouvait déborder au bout du calendrier. Elle applique désormais le même
    plafond, au même endroit du raisonnement.
    """
    debut = date.fromisoformat(payload["range"]["min"])
    fin = date.fromisoformat(payload["range"]["max"])
    if (fin - debut).days > geo.PLAGE_MAX_JOURS:
        fin = debut + timedelta(days=geo.PLAGE_MAX_JOURS)
    if fin < debut:
        fin = debut
    jours = max(1, (fin - debut).days + 1)
    pas = 1
    for candidat in (1, 7, 14, 30):
        pas = candidat
        if jours / candidat <= GRILLE_MAX:
            break
    colonnes = []
    curseur = debut
    while curseur <= fin and len(colonnes) < GRILLE_MAX:
        prochaine = curseur + timedelta(days=pas)
        colonnes.append((curseur, min(prochaine - timedelta(days=1), fin)))
        curseur = prochaine
    return {"pas": pas, "colonnes": colonnes, "debut": debut, "fin": fin}


def _entete_grille(feuille, livre, grille, ligne, col0, accent):
    f = livre.add_format({
        "font_size": 7, "rotation": 90, "align": "center", "valign": "bottom",
        "font_color": "#6B7379", "font_name": "Lexend"})
    for i, (d, _f) in enumerate(grille["colonnes"]):
        feuille.set_column(col0 + i, col0 + i, 3)
        libelle = d.strftime("%Y-%m-%d") if grille["pas"] == 1 else \
            "%d %s" % (d.day, geo.MOIS_FR[d.month - 1])
        feuille.write(ligne, col0 + i, libelle, f)


def _barre_grille(feuille, livre, grille, r, col0, tache, debut, fin):
    fond, plein = geo.COULEURS.get(tache["status"], geo.COULEURS["upcoming"])
    couleur = plein if tache.get("progress", 0) >= 100 else fond
    f = livre.add_format({"bg_color": couleur, "border": 1,
                          "border_color": "#FFFFFF"})
    for i, (cd, cf) in enumerate(grille["colonnes"]):
        if debut <= cf and fin >= cd:
            feuille.write_blank(r, col0 + i, None, f)


# ------------------------------------------------------------------- import

def lire(contenu):
    """Relit un classeur produit par `rendre` (ou taillé pareil).

    Rend une liste de dictionnaires bruts. La conversion en enregistrements est
    le travail de l'assistant, pas celui d'ici : ce module ne connaît pas Odoo.
    """
    from openpyxl import load_workbook

    if len(contenu) > OCTETS_MAX:
        raise ValueError(
            "Classeur trop volumineux : %.1f Mo, la limite est de %d Mo."
            % (len(contenu) / 1048576.0, OCTETS_MAX // 1048576))

    livre = load_workbook(io.BytesIO(contenu), data_only=True, read_only=True)
    feuille = livre.active

    entetes = {}
    ligne_entete = None
    connus = {libelle.lower(): cle for cle, libelle, _w in COLONNES}
    for r, rangee in enumerate(feuille.iter_rows(min_row=1, max_row=25,
                                                 values_only=True), start=1):
        trouve = {}
        for col, valeur in enumerate(rangee):
            if isinstance(valeur, str) and valeur.strip().lower() in connus:
                trouve[connus[valeur.strip().lower()]] = col
        if "name" in trouve and ("start" in trouve or "end" in trouve):
            entetes, ligne_entete = trouve, r
            break

    if not ligne_entete:
        raise ValueError(
            "Aucune ligne d'en-tête reconnue. Le classeur doit porter au moins "
            "les colonnes « Ligne » et « Début »."
        )

    lignes = []
    for rangee in feuille.iter_rows(min_row=ligne_entete + 1,
                                    max_row=ligne_entete + LIGNES_MAX,
                                    values_only=True):
        brut = {cle: (rangee[col] if col < len(rangee) else None)
                for cle, col in entetes.items()}
        nom = (brut.get("name") or "")
        nom = nom.strip() if isinstance(nom, str) else str(nom or "").strip()
        if not nom:
            continue
        if not brut.get("start") and not brut.get("end"):
            # C'est un bandeau de couloir, pas une ligne.
            continue
        lignes.append({
            "name": nom,
            "lane": _texte(brut.get("lane")),
            "assignee": _texte(brut.get("assignee")),
            "start": _date(brut.get("start")),
            "end": _date(brut.get("end")),
            "progress": _entier(brut.get("progress")),
            "status": STATUT_DEPUIS_LIBELLE.get(
                _texte(brut.get("status")).lower(), ""),
            "is_milestone": _texte(brut.get("milestone")).lower() in
                            ("oui", "yes", "x", "vrai", "true", "1"),
            "allocated_hours": _flottant(brut.get("allocated_hours")),
            "depends_on": [p.strip() for p in
                           _texte(brut.get("depends_on")).split(";") if p.strip()],
        })
    livre.close()
    return lignes


def _texte(valeur):
    if valeur is None:
        return ""
    return str(valeur).strip()


def _date(valeur):
    if valeur in (None, ""):
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    texte = str(valeur).strip()[:10]
    try:
        return date.fromisoformat(texte)
    except ValueError:
        for gabarit in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(str(valeur).strip(), gabarit).date()
            except ValueError:
                continue
    return None


def _entier(valeur):
    try:
        return max(0, min(100, int(round(float(valeur)))))
    except (TypeError, ValueError):
        return 0


def _flottant(valeur):
    try:
        return float(valeur)
    except (TypeError, ValueError):
        return 0.0
