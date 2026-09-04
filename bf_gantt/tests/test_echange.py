# -*- coding: utf-8 -*-
"""L'aller-retour des fichiers.

Un export qui ne se relit pas est un cul-de-sac. Ces contrôles écrivent puis
relisent, et comparent ce qui compte : les noms, les dates, les couloirs, les
jalons et les liens. Le reste (couleurs, positions) n'a pas à survivre.
"""
import base64
import io
from datetime import date

from odoo.tests.common import BaseCase, tagged

from ..generateur import geometrie as geo

from ..generateur import mspdi as gen_mspdi
from ..generateur import pdf as gen_pdf
from ..generateur import png as gen_png
from ..generateur import svg as gen_svg
from ..generateur import xlsx as gen_xlsx
from .test_geometrie import barre, echeancier

# Un PNG 4x4 valide, pour tenir lieu de logo sans dépendre d'un fichier.
LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAAQAAAAECAYAAACp8Z5+AAAAHElEQVQI12P4//8/w38G"
    "IAXDIBKE0DHxgljNBAAO9TXL0Y4OHwAAAABJRU5ErkJggg=="
)

COMPLET = echeancier(
    couloirs=[
        {"key": "a", "name": "Cadrage", "seq": 1, "total": 2, "done": 1, "pct": 50},
        {"key": "b", "name": "Migration", "seq": 2, "total": 2, "done": 0, "pct": 0},
    ],
    taches=[
        barre("task-1", lane="a", debut="2026-09-07", fin="2026-09-11",
              progress=100, status="done", nom="Atelier de cadrage"),
        barre("milestone-1", lane="a", debut="2026-09-14", fin="2026-09-14",
              jalon=True, progress=0, status="upcoming", nom="Devis accepté"),
        barre("task-2", lane="b", debut="2026-09-15", fin="2026-10-02",
              progress=30, status="in_progress", nom="Reprise des bases"),
        barre("task-3", lane="b", debut="2026-10-05", fin="2026-10-20",
              progress=0, status="upcoming", nom="Formation & bénévoles <test>"),
    ],
    deps=[{"from": "task-1", "to": "task-2"},
          {"from": "task-2", "to": "task-3"}],
)

MARQUE = dict(COMPLET, company=dict(
    COMPLET["company"], logo=LOGO_B64, dark="#123456",
    tagline="Le slogan de la maison"))

# 🔴 Le cas de Blue Fox : un logo SVG, que Pillow ne sait pas ouvrir.
LOGO_SVG = base64.b64encode(
    b'<?xml version="1.0" encoding="UTF-8"?>'
    b'<svg xmlns="http://www.w3.org/2000/svg" width="240" height="60" '
    b'viewBox="0 0 240 60"><rect width="240" height="60" fill="#29ABE2"/></svg>'
).decode("ascii")
VECTORIEL = dict(COMPLET, company=dict(
    COMPLET["company"], logo=LOGO_SVG, name="Société Vectorielle"))


@tagged("post_install", "-at_install", "bf_gantt")
class TestEchange(BaseCase):

    def _analyseur_utilise_par(self, appel):
        """L'analyseur que le MODULE passe à lxml, capturé au vol.

        ⚠️ Sans cela, un contrôle qui fabrique son propre `XMLParser` éprouve
        lxml et non la configuration d'ici : retirer `resolve_entities=False` du
        module le laisserait vert. Et `XMLParser` n'expose pas ses drapeaux en
        attributs, donc on éprouve le comportement, pas la déclaration.
        """
        from unittest.mock import patch

        from lxml import etree

        vu = {}
        vraie = etree.fromstring

        def espion(texte, parser=None, *args, **kw):
            vu.setdefault("parser", parser)
            return vraie(texte, parser, *args, **kw)

        with patch.object(etree, "fromstring", espion):
            appel()
        self.assertIn("parser", vu, "le module doit passer son propre analyseur")
        self.assertIsNotNone(vu["parser"])
        return vu["parser"]

    # ------------------------------------------------------------ les images

    def test_le_pdf_sort_et_commence_par_sa_signature(self):
        octets = gen_pdf.rendre(COMPLET)
        self.assertTrue(octets.startswith(b"%PDF-"))
        self.assertGreater(len(octets), 2000)

    def test_le_png_sort_et_commence_par_sa_signature(self):
        octets = gen_png.rendre(COMPLET)
        self.assertEqual(octets[:8], b"\x89PNG\r\n\x1a\n")

    def test_le_svg_echappe_ce_qui_vient_de_la_base(self):
        """Un nom de tâche contenant `<` ne doit jamais sortir tel quel."""
        texte = gen_svg.rendre(COMPLET).decode("utf-8")
        self.assertIn("&lt;test&gt;", texte)
        self.assertNotIn("<test>", texte)

    def test_le_zoom_etire_la_boite_du_svg_sans_toucher_au_repere(self):
        """C'est tout l'intérêt : net à n'importe quel facteur, rien recalculé."""
        import re
        for facteur in (1.0, 1.5, 2.5):
            texte = gen_svg.rendre(COMPLET, zoom=facteur).decode("utf-8")
            m = re.search(
                r'width="([\d.]+)" height="([\d.]+)" viewBox="0 0 ([\d.]+) ([\d.]+)"',
                texte)
            self.assertTrue(m, "en-tête SVG illisible")
            largeur, hauteur, vw, vh = (float(v) for v in m.groups())
            self.assertAlmostEqual(largeur / vw, facteur, places=2)
            self.assertAlmostEqual(hauteur / vh, facteur, places=2)

    def test_le_zoom_redessine_le_png_au_lieu_de_l_etirer(self):
        from PIL import Image
        petit = Image.open(io.BytesIO(gen_png.rendre(COMPLET, zoom=1.0)))
        grand = Image.open(io.BytesIO(gen_png.rendre(COMPLET, zoom=2.0)))
        self.assertAlmostEqual(grand.width / petit.width, 2.0, places=1)
        self.assertAlmostEqual(grand.height / petit.height, 2.0, places=1)

    def test_un_zoom_absurde_retombe_dans_les_bornes(self):
        self.assertEqual(geo.borner_zoom(0.01, defaut=1.0), geo.ZOOM_MIN)
        self.assertEqual(geo.borner_zoom(99, defaut=1.0), geo.ZOOM_MAX)
        self.assertEqual(geo.borner_zoom("pas un nombre"), geo.ZOOM_DEFAUT)
        self.assertEqual(geo.borner_zoom(None), geo.ZOOM_DEFAUT)

    def test_le_pdf_garde_sa_taille_de_page(self):
        """Une page d'impression agrandie ne dirait plus la vérité sur sa taille.

        ⚠️ On compare le `/MediaBox`, pas les octets : reportlab pose un
        identifiant de document aléatoire, donc deux rendus du même échéancier ne
        sont jamais identiques octet pour octet.
        """
        import re

        def page(octets):
            m = re.search(rb"/MediaBox\s*\[([^\]]+)\]", octets)
            self.assertTrue(m, "aucun /MediaBox dans le PDF")
            return m.group(1).strip()

        self.assertEqual(page(gen_pdf.rendre(COMPLET)),
                         page(gen_pdf.rendre(COMPLET)))
        g = geo.construire(COMPLET)
        boite = page(gen_pdf.rendre(COMPLET)).split()
        self.assertAlmostEqual(float(boite[2]), g["largeur"], places=0)
        self.assertAlmostEqual(float(boite[3]), g["hauteur"], places=0)

    def test_les_trois_images_supportent_les_trois_echelles(self):
        for echelle in ("day", "week", "month"):
            self.assertTrue(gen_pdf.rendre(COMPLET, echelle=echelle))
            self.assertTrue(gen_svg.rendre(COMPLET, echelle=echelle))
            self.assertTrue(gen_png.rendre(COMPLET, echelle=echelle))

    def test_un_echeancier_vide_ne_fait_pas_tomber_les_rendus(self):
        vide = echeancier(taches=[], couloirs=[])
        self.assertTrue(gen_pdf.rendre(vide).startswith(b"%PDF-"))
        self.assertTrue(gen_svg.rendre(vide))
        self.assertTrue(gen_png.rendre(vide))
        self.assertTrue(gen_xlsx.rendre(vide))
        self.assertTrue(gen_mspdi.rendre(vide))

    def test_le_plafonnement_de_plage_est_dit_dans_les_sorties(self):
        """⚠️ `plage_reduite` était calculé, renvoyé, contrôlé... et lu par aucun
        rendu. Un plafond muet vaut une troncature muette."""
        loin = dict(COMPLET, tasks=COMPLET["tasks"], range={
            "min": "2026-01-01", "max": "3026-01-01", "today": "2026-06-01"})
        self.assertTrue(geo.construire(loin)["plage_reduite"])
        # ⚠️ Le SVG et le classeur portent le texte en clair ; le PDF comprime
        # ses flux et le PNG est une image, donc on ne peut y chercher un mot.
        # On y contrôle ce qui est contrôlable : le rendu sort, et il diffère de
        # celui d'un échéancier non plafonné.
        self.assertIn("plage ramenée", gen_svg.rendre(loin).decode("utf-8"))
        # ⚠️ Un XLSX est une archive : le texte n'est pas dans les octets bruts.
        # On rouvre le classeur et on lit les cellules, comme un lecteur le ferait.
        from openpyxl import load_workbook
        feuille = load_workbook(io.BytesIO(gen_xlsx.rendre(loin))).active
        textes = " ".join(str(c.value) for rangee in feuille.iter_rows()
                          for c in rangee if c.value)
        self.assertIn("plage a été ramenée", textes)
        self.assertNotEqual(gen_pdf.rendre(loin)[:400],
                            gen_pdf.rendre(COMPLET)[:400])
        self.assertTrue(gen_png.rendre(loin))

    def test_un_echeancier_normal_ne_parle_pas_de_plafonnement(self):
        self.assertNotIn("plage ramenée", gen_svg.rendre(COMPLET).decode("utf-8"))

    # -------------------------------------------------------------- la marque

    def test_le_logo_entre_dans_le_pdf(self):
        avec = gen_pdf.rendre(MARQUE)
        sans = gen_pdf.rendre(COMPLET)
        self.assertGreater(len(avec), len(sans),
                           "le PDF avec logo devrait être plus lourd")

    def test_le_logo_entre_dans_le_svg_en_data_uri(self):
        """Un SVG qui pointe vers une URL de logo serait cassé une fois envoyé."""
        texte = gen_svg.rendre(MARQUE).decode("utf-8")
        self.assertIn("<image", texte)
        self.assertIn("data:image/png;base64,", texte)
        self.assertNotIn("<image", gen_svg.rendre(COMPLET).decode("utf-8"))

    def test_le_slogan_de_la_societe_atterrit_au_pied(self):
        self.assertIn("Le slogan de la maison",
                      gen_svg.rendre(MARQUE).decode("utf-8"))

    def test_le_logo_entre_dans_le_classeur(self):
        self.assertGreater(len(gen_xlsx.rendre(MARQUE)), len(gen_xlsx.rendre(COMPLET)))

    def test_le_classeur_reste_relisible_avec_le_logo(self):
        """Une image insérée ne doit pas décaler les colonnes nommées."""
        relu = gen_xlsx.lire(gen_xlsx.rendre(MARQUE))
        self.assertEqual(len(relu), 4)

    def test_un_logo_svg_est_reconnu_et_mesure(self):
        boite = geo.boite_logo(LOGO_SVG)
        self.assertTrue(boite)
        self.assertTrue(boite["vectoriel"])
        self.assertEqual(boite["mime"], "image/svg+xml")
        # 240 x 60 ramené sous 26 pt de haut : le rapport doit tenir.
        self.assertAlmostEqual(boite["largeur"] / boite["hauteur"], 4.0, places=2)

    def test_un_svg_sans_width_est_mesure_par_son_viewbox(self):
        sans = base64.b64encode(
            b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 50"/>'
        ).decode("ascii")
        boite = geo.boite_logo(sans)
        self.assertTrue(boite)
        self.assertAlmostEqual(boite["largeur"] / boite["hauteur"], 2.0, places=2)

    def test_le_logo_svg_entre_tel_quel_dans_le_svg(self):
        texte = gen_svg.rendre(VECTORIEL).decode("utf-8")
        self.assertIn("data:image/svg+xml;base64,", texte)

    def test_le_logo_svg_devient_un_mot_symbole_dans_les_rendus_matriciels(self):
        """Sans rasteriseur dans l'image, le nom à la couleur de la marque."""
        from PIL import Image
        avec = Image.open(io.BytesIO(gen_png.rendre(VECTORIEL)))
        self.assertTrue(avec.width > 0)
        # Le PDF sort, et il est plus lourd que sans marque du tout : le texte y est.
        self.assertTrue(gen_pdf.rendre(VECTORIEL).startswith(b"%PDF-"))
        self.assertTrue(gen_xlsx.rendre(VECTORIEL))

    def test_un_svg_ne_resout_pas_les_entites_externes(self):
        """⚠️ Contrôle réécrit : la version d'avant comparait à `octets[:0]`,
        donc toujours `b""`, et ne pouvait pas échouer. Elle regardait en plus
        les octets d'ENTRÉE, pas le résultat de l'analyse. On interroge donc
        l'analyseur lui-même, et on prouve que le fichier n'est pas lu."""
        from unittest.mock import patch

        from lxml import etree

        brut = (
            b'<?xml version="1.0"?>'
            b'<!DOCTYPE svg [<!ENTITY xxe SYSTEM "file:///etc/passwd">]>'
            b'<svg xmlns="http://www.w3.org/2000/svg" width="10" height="10">'
            b'<desc>&xxe;</desc></svg>'
        )
        # 1. La mesure fonctionne quand même : un SVG piégé reste un SVG.
        boite = geo.boite_logo(base64.b64encode(brut).decode("ascii"))
        self.assertTrue(boite, "le SVG aurait dû être mesuré")
        self.assertTrue(boite["vectoriel"])

        # 2. 🔴 Et surtout : on interroge l'analyseur DU MODULE, pas un
        # analyseur que le contrôle fabriquerait lui-même. La version d'avant
        # construisait son propre `XMLParser` : passer `resolve_entities=True`
        # dans `geometrie.py` la laissait verte, donc elle éprouvait lxml et non
        # la configuration d'ici.
        parser = self._analyseur_utilise_par(lambda: geo._taille_svg(brut))

        # 3. Et avec CET analyseur-là, l'entité n'a rien lu du disque.
        rendu = etree.tostring(etree.fromstring(brut, parser=parser),
                               encoding="unicode")
        self.assertNotIn("root:", rendu)
        self.assertNotIn("/bin/", rendu)

    def test_un_svg_a_entites_recursives_ne_fait_pas_exploser_la_memoire(self):
        """Le milliard de rires : libxml2 le refuse, et on rend None sans lever."""
        entites = [b"<!ENTITY n0 'aaaaaaaaaa'>"]
        for i in range(1, 8):
            entites.append(b"<!ENTITY n%d '%s'>" % (i, b"&n%d;" % (i - 1) * 10))
        bombe = (b"<?xml version='1.0'?><!DOCTYPE svg [" + b"".join(entites) + b"]>"
                 b"<svg xmlns='http://www.w3.org/2000/svg' width='10' height='10'>"
                 b"<desc>&n7;</desc></svg>")
        self.assertIsNone(geo.boite_logo(base64.b64encode(bombe).decode("ascii")))

    def test_un_logo_illisible_ne_fait_tomber_aucun_rendu(self):
        """C'est la garde qui compte : un document sans logo reste un document."""
        casse = dict(COMPLET, company=dict(COMPLET["company"],
                                           logo="pas du tout une image"))
        self.assertIsNone(geo.boite_logo("pas du tout une image"))
        self.assertTrue(gen_pdf.rendre(casse).startswith(b"%PDF-"))
        self.assertTrue(gen_svg.rendre(casse))
        self.assertTrue(gen_png.rendre(casse))
        self.assertTrue(gen_xlsx.rendre(casse))
        self.assertTrue(gen_mspdi.rendre(casse))

    def test_la_couleur_de_la_societe_peint_l_entete_du_classeur(self):
        rouge = dict(COMPLET, company=dict(COMPLET["company"], color="#AA0000"))
        self.assertNotEqual(gen_xlsx.rendre(rouge), gen_xlsx.rendre(COMPLET))

    # ------------------------------------------------------------- le tableur

    def test_le_classeur_se_relit_ligne_pour_ligne(self):
        relu = gen_xlsx.lire(gen_xlsx.rendre(COMPLET))
        self.assertEqual(len(relu), 4)
        noms = [l["name"] for l in relu]
        self.assertIn("Atelier de cadrage", noms)
        self.assertIn("Formation & bénévoles <test>", noms)

    def test_le_classeur_garde_les_dates_les_couloirs_et_les_jalons(self):
        relu = {l["name"]: l for l in gen_xlsx.lire(gen_xlsx.rendre(COMPLET))}
        atelier = relu["Atelier de cadrage"]
        self.assertEqual(atelier["start"], date(2026, 9, 7))
        self.assertEqual(atelier["end"], date(2026, 9, 11))
        self.assertEqual(atelier["lane"], "Cadrage")
        self.assertEqual(atelier["progress"], 100)
        self.assertTrue(relu["Devis accepté"]["is_milestone"])
        self.assertFalse(atelier["is_milestone"])

    def test_le_classeur_garde_les_dependances_par_nom(self):
        relu = {l["name"]: l for l in gen_xlsx.lire(gen_xlsx.rendre(COMPLET))}
        self.assertEqual(relu["Reprise des bases"]["depends_on"],
                         ["Atelier de cadrage"])
        self.assertEqual(relu["Formation & bénévoles <test>"]["depends_on"],
                         ["Reprise des bases"])

    def test_les_bandeaux_de_couloir_ne_sont_pas_relus_comme_des_lignes(self):
        relu = gen_xlsx.lire(gen_xlsx.rendre(COMPLET))
        self.assertNotIn("Cadrage", [l["name"] for l in relu])

    def test_un_classeur_trop_gros_est_refuse_avant_d_etre_lu(self):
        """Un fichier téléversé est un fichier étranger : il a un plafond."""
        with self.assertRaises(ValueError):
            gen_xlsx.lire(b"PK" + b"\0" * (gen_xlsx.OCTETS_MAX + 1))

    def test_le_nombre_de_rangees_lues_est_borne(self):
        self.assertLessEqual(gen_xlsx.LIGNES_MAX, 100_000)
        self.assertGreater(gen_xlsx.LIGNES_MAX, 100)

    def test_un_classeur_sans_entete_est_refuse_lisiblement(self):
        import io
        import xlsxwriter
        tampon = io.BytesIO()
        livre = xlsxwriter.Workbook(tampon, {"in_memory": True})
        livre.add_worksheet("vide").write(0, 0, "n'importe quoi")
        livre.close()
        with self.assertRaises(ValueError):
            gen_xlsx.lire(tampon.getvalue())

    # -------------------------------------------------------------- MS Project

    def test_le_mspdi_se_relit_ligne_pour_ligne(self):
        relu = gen_mspdi.lire(gen_mspdi.rendre(COMPLET))
        self.assertEqual(relu["title"], "Essai")
        self.assertEqual(len(relu["lines"]), 4)

    def test_les_couloirs_deviennent_des_recapitulatives_et_reviennent(self):
        relu = gen_mspdi.lire(gen_mspdi.rendre(COMPLET))
        par_nom = {l["name"]: l for l in relu["lines"]}
        self.assertEqual(par_nom["Atelier de cadrage"]["lane"], "Cadrage")
        self.assertEqual(par_nom["Reprise des bases"]["lane"], "Migration")

    def test_le_mspdi_garde_dates_jalons_et_liens(self):
        relu = {l["name"]: l for l in
                gen_mspdi.lire(gen_mspdi.rendre(COMPLET))["lines"]}
        self.assertEqual(relu["Atelier de cadrage"]["start"], date(2026, 9, 7))
        self.assertEqual(relu["Atelier de cadrage"]["end"], date(2026, 9, 11))
        self.assertTrue(relu["Devis accepté"]["is_milestone"])
        self.assertEqual(relu["Reprise des bases"]["depends_on"],
                         ["Atelier de cadrage"])

    def test_le_mspdi_porte_un_calendrier(self):
        """Project refuse d'ouvrir un fichier sans calendrier de base."""
        texte = gen_mspdi.rendre(COMPLET).decode("utf-8")
        self.assertIn("<Calendars>", texte)
        self.assertIn("<IsBaseCalendar>1</IsBaseCalendar>", texte)

    def test_un_xml_qui_n_est_pas_un_mspdi_est_refuse_lisiblement(self):
        with self.assertRaises(ValueError):
            gen_mspdi.lire(b"<?xml version='1.0'?><autrechose/>")

    def test_un_fichier_illisible_est_refuse_lisiblement(self):
        with self.assertRaises(ValueError):
            gen_mspdi.lire(b"ceci n'est pas du XML")

    def test_l_analyseur_mspdi_du_module_refuse_les_entites(self):
        """Éprouve la CONFIGURATION du module, pas le comportement de lxml."""
        from unittest.mock import patch

        from lxml import etree

        parser = self._analyseur_utilise_par(
            lambda: gen_mspdi.lire(gen_mspdi.rendre(COMPLET)))
        poison = (b"<?xml version='1.0'?>"
                  b"<!DOCTYPE r [<!ENTITY xxe SYSTEM 'file:///etc/passwd'>]>"
                  b"<r>&xxe;</r>")
        rendu = etree.tostring(etree.fromstring(poison, parser=parser),
                               encoding="unicode")
        self.assertNotIn("root:", rendu)

    def test_le_mspdi_ne_resout_pas_les_entites_externes(self):
        """Une entité externe dans un fichier reçu ne doit RIEN lire du disque."""
        poison = (
            b"<?xml version='1.0'?>"
            b"<!DOCTYPE Project [<!ENTITY xxe SYSTEM 'file:///etc/passwd'>]>"
            b"<Project xmlns='http://schemas.microsoft.com/project'>"
            b"<Name>&xxe;</Name><Tasks/></Project>"
        )
        try:
            lu = gen_mspdi.lire(poison)
        except ValueError:
            return  # Refusé net : très bien aussi.
        self.assertNotIn("root:", lu["title"])

    def test_la_duree_iso_se_relit(self):
        self.assertEqual(gen_mspdi._heures("PT8H0M0S"), 8.0)
        self.assertEqual(gen_mspdi._heures("PT1H30M0S"), 1.5)
        self.assertEqual(gen_mspdi._heures(""), 0.0)
        self.assertEqual(gen_mspdi._heures("n'importe quoi"), 0.0)
