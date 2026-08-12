# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).
"""Wizard d'import xlsx/ods pour les pools de licences.

Lit un fichier au format colonnes ``Date d'achat | Clé de produit |
Activé le | Client | Commentaires/notes`` (le format BF existant pour
Windows 11 Pro / Office 2024) et crée un ``hosting.license`` + autant de
``hosting.license.seat`` qu'il y a de lignes.
"""
import base64
import io
import logging
from datetime import date, datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


def _normalize_str(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _parse_date(value):
    if not value:
        return False
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return False


class HostingLicenseImportWizard(models.TransientModel):
    _name = "hosting.license.import.wizard"
    _description = "Import d'un pool de licences (xlsx/ods)"

    attachment = fields.Binary(string="Fichier xlsx/ods", required=True)
    filename = fields.Char(string="Nom du fichier")

    name = fields.Char(
        string="Nom du pool",
        required=True,
        help="Ex. : « Windows 11 Pro – batch 2025-09 »",
    )
    software_id = fields.Many2one(
        comodel_name="hosting.software",
        string="Logiciel (catalogue)",
    )
    product_name = fields.Char(string="Produit (texte libre)")
    vendor = fields.Char(string="Éditeur", default="Microsoft")
    license_type = fields.Selection(
        selection=[
            ("per_device", "Une clé par poste"),
            ("per_user", "Une clé par utilisateur"),
            ("volume", "Clé volume"),
            ("subscription", "Abonnement"),
        ],
        string="Type",
        default="per_device",
        required=True,
    )
    owner_partner_id = fields.Many2one(
        comodel_name="res.partner",
        string="Propriétaire",
        required=True,
        default=lambda self: self.env.company.partner_id,
    )
    purchase_date = fields.Date(
        string="Date d'achat (défaut)",
        help="Utilisée si la colonne « Date d'achat » est vide dans le fichier.",
    )

    def _read_rows(self):
        """Lit le binary et retourne une liste de listes (chaque ligne = liste
        de cellules). Détecte xlsx vs ods par signature."""
        if not self.attachment:
            raise UserError(_("Aucun fichier fourni."))
        data = base64.b64decode(self.attachment)
        # xlsx & ods sont tous deux des ZIPs (signature PK\x03\x04)
        # On différencie par le contenu : xlsx contient xl/, ods contient content.xml.
        import zipfile

        try:
            zf = zipfile.ZipFile(io.BytesIO(data))
        except zipfile.BadZipFile as exc:
            raise UserError(_("Fichier illisible (ni xlsx ni ods).")) from exc
        names = set(zf.namelist())
        if "xl/workbook.xml" in names:
            return self._read_xlsx(data)
        if "content.xml" in names:
            return self._read_ods(zf)
        raise UserError(_("Format non reconnu (xlsx ou ods attendu)."))

    def _read_xlsx(self, data):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UserError(
                _("openpyxl est requis (devrait être disponible par défaut dans Odoo).")
            ) from exc
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows

    def _read_ods(self, zf):
        import xml.etree.ElementTree as ET

        ns = {
            "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
            "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
            "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
        }
        with zf.open("content.xml") as fp:
            tree = ET.parse(fp)
        root = tree.getroot()
        rows = []
        for trow in root.iter(f"{{{ns['table']}}}table-row"):
            cells = []
            for tcell in trow.iter(f"{{{ns['table']}}}table-cell"):
                text = "".join(tcell.itertext()).strip()
                cells.append(text)
            # Retirer les cellules vides en fin de ligne
            while cells and not cells[-1]:
                cells.pop()
            if cells:
                rows.append(cells)
        return rows

    @staticmethod
    def _find_header_index(rows):
        """Trouve l'index de la ligne d'en-tête (celle contenant « clé de
        produit » ou « clé » / « key »). Retourne -1 si introuvable."""
        for idx, row in enumerate(rows):
            joined = " ".join(str(c or "").strip().lower() for c in row)
            if "clé de produit" in joined or "product key" in joined or (
                "clé" in joined and "achat" in joined
            ):
                return idx
        return -1

    def action_import(self):
        self.ensure_one()
        rows = self._read_rows()
        if not rows:
            raise UserError(_("Le fichier est vide."))

        # Détecter la ligne d'en-tête (les fichiers réels ont des lignes vides
        # ou un titre avant). Les données commencent juste après.
        header_idx = self._find_header_index(rows)
        if header_idx < 0:
            # Pas d'en-tête reconnu : on suppose la 1re ligne non vide = header
            header_idx = next(
                (i for i, r in enumerate(rows) if any(str(c or "").strip() for c in r)),
                0,
            )
        data_rows = rows[header_idx + 1:]

        # Schéma attendu : 5 colonnes positionnelles
        # Date d'achat | Clé de produit | Activé le | Client | Notes
        License = self.env["hosting.license"]
        Seat = self.env["hosting.license.seat"]

        license_vals = {
            "name": self.name,
            "software_id": self.software_id.id if self.software_id else False,
            "product_name": self.product_name or False,
            "vendor": self.vendor or False,
            "license_type": self.license_type,
            "owner_partner_id": self.owner_partner_id.id,
            "purchase_date": self.purchase_date or False,
        }
        license_rec = License.create(license_vals)

        created_seats = 0
        for row in data_rows:
            cells = (list(row) + [None] * 5)[:5]
            purchase_d, key, activated_d, client, notes = cells
            key = _normalize_str(key)
            if not key:
                continue
            seat_vals = {
                "license_id": license_rec.id,
                "product_key": key,
                "notes": _normalize_str(notes) or False,
                "activated_on": _parse_date(activated_d) or False,
            }
            client_str = _normalize_str(client)
            if client_str:
                seat_vals["assignee_type"] = "free_text"
                seat_vals["assignee_label"] = client_str
            # State : activated si « Activé le » non vide, failed si notes contiennent "DOESN'T WORK"
            notes_str = _normalize_str(notes).lower()
            if "doesn" in notes_str or "n'a pas marché" in notes_str:
                seat_vals["state"] = "failed"
            elif seat_vals.get("activated_on"):
                seat_vals["state"] = "activated"
            else:
                seat_vals["state"] = "free"
            Seat.create(seat_vals)
            created_seats += 1

        # Mettre seats_total à au moins le nombre de sièges importés
        if created_seats > license_rec.seats_total:
            license_rec.seats_total = created_seats

        if not created_seats:
            raise UserError(
                _("Aucune clé trouvée. Vérifiez que le fichier a une colonne "
                  "« Clé de produit » et des lignes de données.")
            )

        # Date d'achat globale : prendre la 1re non vide si pas saisie au wizard
        if not license_rec.purchase_date:
            for row in data_rows:
                d = _parse_date((list(row) + [None])[0])
                if d:
                    license_rec.purchase_date = d
                    break

        return {
            "type": "ir.actions.act_window",
            "name": _("Pool importé"),
            "res_model": "hosting.license",
            "res_id": license_rec.id,
            "view_mode": "form",
            "target": "current",
        }
